import openpyxl
from openpyxl.styles import PatternFill, Border, Side

def insert_blank_rows(save_path):
    wb = openpyxl.load_workbook(save_path)
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ws in wb.worksheets:
        if ws.title == '전체':
            continue
        header = [cell.value for cell in ws[1]]
        if '요일' not in header:
            continue
        day_col_idx = header.index('요일') + 1

        for row in range(ws.max_row, 1, -1):
            cell_value = ws.cell(row=row, column=day_col_idx).value
            if cell_value == '일':
                ws.insert_rows(row + 1)
                for col in range(1, 18):
                    cell = ws.cell(row=row + 1, column=col)
                    cell.fill = yellow_fill
                    cell.border = border

        # ✅ 마지막 데이터 행 아래에 노란색 빈 행 삽입
        last_row = ws.max_row + 1
        ws.insert_rows(last_row)
        for col in range(1, 18):
            cell = ws.cell(row=last_row, column=col)
            cell.fill = yellow_fill
            cell.border = border

    wb.save(save_path)
    wb.close()