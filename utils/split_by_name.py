import openpyxl
from collections import OrderedDict

def split_by_name(save_path):
    wb = openpyxl.load_workbook(save_path)
    ws = wb.active

    # 1. 기존 시트 이름을 '전체'로 변경
    ws.title = '전체'

    # 2. 헤더 추출
    header = [cell.value for cell in ws[1]]
    name_col_idx = header.index('이름')  # '이름' 열 위치

    # 3. 이름별로 행 분류
    name_rows = OrderedDict()

    for row in ws.iter_rows(min_row=2, values_only=False):
        name = row[name_col_idx].value
        if name not in name_rows:
            name_rows[name] = []
        name_rows[name].append(row)

    # 4. 이름별 시트 생성
    for name, rows in name_rows.items():
        new_ws = wb.create_sheet(title=str(name))

        # 헤더 복사
        for col_idx, cell in enumerate(ws[1], start=1):
            new_ws.cell(row=1, column=col_idx, value=cell.value)

        # 데이터 복사 (수식 포함)
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, cell in enumerate(row, start=1):
                new_ws.cell(row=row_idx, column=col_idx, value=cell.value)

    wb.save(save_path)
    wb.close()