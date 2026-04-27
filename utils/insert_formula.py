import openpyxl
from openpyxl.utils import get_column_letter

def insert_formula(save_path):
    wb = openpyxl.load_workbook(save_path)

    for ws in wb.worksheets:
        if ws.title == '전체':
            continue
        # 열 이름으로 열 번호 찾기
        header = [cell.value for cell in ws[1]]

        try:
            col_K = header.index('근무시간(분)')+1
            col_O = header.index('저녁시간(분)')+1
            col_P = header.index('출근시간전(분)')+1
            col_result = header.index('실제근무시간(K-O-Q)')+1
        except ValueError as e:
            print(f"[{ws.title}] 열을 찾을 수 없습니다: {e}")
            continue        

        k = get_column_letter(col_K)
        o = get_column_letter(col_O)
        p = get_column_letter(col_P)
        r = get_column_letter(col_result)

        # 2행부터 마지막 행까지 수식 삽입
        for row in range(2, ws.max_row + 1):
            # 휴일 행(근무시간이 빈 셀)은 수식 삽입 안 함
            if ws[f'{k}{row}'].value in (None, ''):
                continue
            ws[f'{r}{row}'] = f'={k}{row}-({o}{row}+{p}{row})'
        #틀고정
        ws.freeze_panes = 'K2'

    wb.save(save_path)
    wb.close()