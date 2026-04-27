import openpyxl
from openpyxl.styles import Border, Side

def apply_border(save_path, skip_title='전체'):
    wb = openpyxl.load_workbook(save_path)
    
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ws in wb.worksheets:
        if skip_title and ws.title == skip_title:
            continue

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, 
                                 min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = border

    wb.save(save_path)
    wb.close()