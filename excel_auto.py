import pandas as pd
import os
import customtkinter as ctk
import openpyxl

from openpyxl.utils import get_column_letter
from tkinter import filedialog, messagebox
from collections import OrderedDict

selected_file_path = ""


def select_file():
    global selected_file_path
    path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
    if path:
        selected_file_path = path
        label.configure(text=f"선택됨: {os.path.basename(path)}")

        # #파일명만 추출해서 라벨에 표시
        # filename = os.path.basename(path)
        # label.configure(text=f"선택된 파일: {filename}")
        # print(f"선택된 경로: {selected_file_path}")

def delete_column(df):
    # 1. 열 삭제
    cols_to_delete = ['사번', '출근스케줄', '퇴근스케줄', '휴게시간', '근무시간', '제외시간', '연장근무시간(신청)', '야간근무시간(신청)', 
                      '야간근무시간(실제)', '외근-간주시간(신청)', '제외시간(분)', '연장근무시간(신청)(분)', '야간근무시간(신청)(분)',
                      '외근-간주시간(신청)(분)', '출근입력시간', '퇴근입력시간', '자리비움시간(RAW)', '외근-간주시간']
    
    existing_cols = [col for col in cols_to_delete if col in df.columns]
    if existing_cols:
        df = df.drop(columns=existing_cols)
    
    return df

def modify_rows(df):
    #특정 조건 행 삭제(CEO)
    #1. "a"열 값이 CEO가 아닌 데이터만 추출
    if 'Team' in df.columns:
        df=df[df['Team'] != 'CEO']
        
    return df

def add_columns(df):
    #근무시간(시간) 열 뒤에 새로운 열 2개 추가

    target_col='근무시간(분)'

    if target_col in df.columns:
        #1. 대상 열 위치 찾기
        idx = df.columns.get_loc(target_col)

        #2. 열 삽입 (대상 열 바로 뒤이므로 idx+1)
        if '실제근무시간(K-O-Q)' not in df.columns:
            df.insert(idx+1, '실제근무시간(K-O-Q)', '')
        if '실제근무시간(시.분)' not in df.columns:
            df.insert(idx+2, '실제근무시간(시.분)', '')

    else:
        print(f"경고: '{target_col}'열을 찾을수 없습니다.")

    return df

def insert_formula(save_path):
    wb = openpyxl.load_workbook(save_path)

    for ws in wb.worksheets:
        if ws.title == '전체':
            continue
        # 열 이름으로 열 번호 찾기
        header = [cell.value for cell in ws[1]]

        try:
            col_K = header.index('근무시간(분)')+1
            col_O = header.index('저녁시간(분)')+1
            col_P = header.index('출근시간전(분)')+1
            col_result = header.index('실제근무시간(K-O-Q)')+1
        except ValueError as e:
            print(f"[{ws.title}] 열을 찾을 수 없습니다: {e}")
            continue
            # wb.close()
        

        k = get_column_letter(col_K)
        o = get_column_letter(col_O)
        p = get_column_letter(col_P)
        r = get_column_letter(col_result)

        # 2행부터 마지막 행까지 수식 삽입
        for row in range(2, ws.max_row + 1):
            # 휴일 행(근무시간이 빈 셀)은 수식 삽입 안 함
            if ws[f'{k}{row}'].value in (None, ''):
                continue
            ws[f'{r}{row}'] = f'={k}{row}-({o}{row}+{p}{row})'
        #틀고정
        ws.freeze_panes = 'K2'

    wb.save(save_path)
    wb.close()

def split_by_name(save_path):
    wb = openpyxl.load_workbook(save_path)
    ws = wb.active

    # 1. 기존 시트 이름을 '전체'로 변경
    ws.title = '전체'

    # 2. 헤더 추출
    header = [cell.value for cell in ws[1]]
    name_col_idx = header.index('이름')  # '이름' 열 위치

    # 3. 이름별로 행 분류
    name_rows = OrderedDict()

    for row in ws.iter_rows(min_row=2, values_only=False):
        name = row[name_col_idx].value
        if name not in name_rows:
            name_rows[name] = []
        name_rows[name].append(row)

    # 4. 이름별 시트 생성
    for name, rows in name_rows.items():
        new_ws = wb.create_sheet(title=str(name))

        # 헤더 복사
        for col_idx, cell in enumerate(ws[1], start=1):
            new_ws.cell(row=1, column=col_idx, value=cell.value)

        # 데이터 복사 (수식 포함)
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, cell in enumerate(row, start=1):
                new_ws.cell(row=row_idx, column=col_idx, value=cell.value)

    wb.save(save_path)
    wb.close()


def insert_blank_rows(save_path):
    wb = openpyxl.load_workbook(save_path)
    
    for ws in wb.worksheets:
        # '전체' 시트는 건너뜀 (개인 시트에만 적용)
        if ws.title == '전체':
            continue

        # 헤더 추출
        header = [cell.value for cell in ws[1]]
        
        if '요일' not in header:
            continue
            
        day_col_idx = header.index('요일') + 1  # 1-based

        # 아래에서 위로 순회 (행 삽입 시 인덱스 밀림 방지)
        for row in range(ws.max_row, 1, -1):
            cell_value = ws.cell(row=row, column=day_col_idx).value
            if cell_value == '일':
                ws.insert_rows(row + 1)  # '일' 행 바로 아래에 빈 행 삽입

    wb.save(save_path)
    wb.close()

def run_automation():
    global selected_file_path

    if not selected_file_path:
        messagebox.showwarning("경고", "먼저 엑셀 파일을 선택해 주세요")
        return
    
    btn_run.configure(state="disabled")  # 중복 실행 방지
    try:
        #1. 파일 불러오기
        df = pd.read_excel(selected_file_path)
    
        #2. 자동화 로직
        #열삭제
        df = delete_column(df)
        #CEO 행 삭제
        df = modify_rows(df)
        #행 삽입
        df = add_columns(df)
        
        
        #3. 결과 저장
        base, ext = os.path.splitext(selected_file_path)
        if base.endswith("_결과"):
            save_path = f"{base}{ext}"
        else:
            save_path = f"{base}_결과{ext}"
        
        df.to_excel(save_path, index=False)

        #시트 분리 (이름별)
        split_by_name(save_path)
        # 2. 빈 행 삽입 (일요일 기준)
        insert_blank_rows(save_path)
        # 3. 수식 삽입 (저장 후 openpyxl로 후처리)
        insert_formula(save_path)

        label.configure(text=f"작업 완료! 결과 파일 생성됨")
        messagebox.showinfo("성공", f"파일이 성공적으로 저장되었습니다.\n{save_path}")
        
    except Exception as e:
        messagebox.showerror("에러", f"오류가 발생했습니다: {e}")
    finally:
        btn_run.configure(state="normal")

#UI 설정
app = ctk.CTk()
app.title("근태 정리 자동화 프로그램")
app.geometry("500x300")

label = ctk.CTkLabel(app, text="엑셀 파일을 선택하세요", font=("Pretendard", 14))
label.pack(pady=30)

#버튼 1: 파일 선택
btn_select = ctk.CTkButton(app, text="파일 불러오기", command=select_file)
btn_select.pack(pady=10)

#버튼 2: 실행
btn_run = ctk.CTkButton(app, text="자동화 실행", command=run_automation, fg_color="green", hover_color="darkgreen")
btn_run.pack(pady=10)

app.mainloop()